import json
from pathlib import Path

from openpyxl import load_workbook

PATH = Path(__file__).parent

def get_data(excel_file=None):
    wb = load_workbook(excel_file, read_only=True)
    data = {}

    data['Penyakit'] = {
        key.title(): {
            'gejala': [int(i) for i in val.split()],
            'prob': 0
        } for key, val in list(
            wb['Penyakit'].iter_rows(values_only=True)
        )[1:]
    }
    data['Gejala'] = {
        str(key): val.title() for key, val in list(
            wb['Gejala'].iter_rows(values_only=True)
        )[1:]
    }
    data['max_symptoms'] = max(data['Penyakit'].values(), key=len)

    return data

def _json_to_js(data):
    with open(PATH / 'json.js', 'w+') as js:
        js.write(f'const json_data = {data};')

if __name__ == '__main__':
    dict_data = get_data(PATH / 'THT.xlsx')

    with open(PATH / 'data.json', 'w+') as json_file:
        json.dump(dict_data, json_file)

    _json_to_js(dict_data)
